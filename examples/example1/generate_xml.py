"""
Generate all Example 1 outputs from Example1.xlsx:
  - output/define.xml
  - output/rwd-lineage.xml
  - data/csv/sdtm/ce.csv
  - data/csv/source/pt_dx.csv
  - data/csv/source/vitals.csv
  - data/csv/source/notes.csv

Run this script any time Example1.xlsx changes to keep everything in sync.
"""
import csv
import uuid
import xml.dom.minidom as minidom
import xml.etree.ElementTree as ET
from datetime import datetime
from pathlib import Path

import openpyxl

EXCEL_PATH = Path(__file__).parent / "data" / "excel" / "Example1.xlsx"
OUTPUT_DIR = Path(__file__).parent / "output"

# Custom namespace UUID for CDISC RWDL (UUID v5 based on DNS namespace + domain string)
RWDL_NS = uuid.uuid5(uuid.NAMESPACE_DNS, "cdisc.org/rwd-lineage")

DEFINE_NS   = "http://www.cdisc.org/ns/odm/v1.3"
DEFINE_NS2  = "http://www.cdisc.org/ns/def/v2.1"
RWDL_NS_URI = "http://www.cdisc.org/ns/rwd-lineage/v1"
XL_NS       = "http://www.w3.org/1999/xlink"
XSI_NS      = "http://www.w3.org/2001/XMLSchema-instance"


def make_uuid(source_coords: str, target_coords: str) -> str:
    key = f"{source_coords}|{target_coords}"
    return str(uuid.uuid5(RWDL_NS, key))


def pretty_xml(element: ET.Element) -> str:
    rough = ET.tostring(element, encoding="unicode")
    dom = minidom.parseString(rough)
    return dom.toprettyxml(indent="  ", encoding=None)


def load_lineage_rows() -> list[dict]:
    """Read data rows from RWDLineage-Table, skipping the two header rows."""
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb["RWDLineage-Table"]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 2:
            continue  # skip group header + column header rows
        tgt_storage, tgt_structure, tgt_uri, tgt_row, tgt_col, \
        src_storage, src_structure, src_uri, src_row, src_col, \
        span_begin, span_end, value = row[1:]
        if tgt_uri is None:
            continue  # skip empty rows
        rows.append({
            "tgt_storage":   tgt_storage,
            "tgt_structure": tgt_structure,
            "tgt_uri":       tgt_uri,
            "tgt_row":       int(tgt_row),
            "tgt_col":       tgt_col,
            "src_storage":   src_storage,
            "src_structure": src_structure,
            "src_uri":       src_uri,
            "src_row":       int(src_row),
            "src_col":       src_col,
            "span_begin":    int(span_begin) if span_begin is not None else None,
            "span_end":      int(span_end)   if span_end   is not None else None,
            "value":         value,
        })
    return rows


def collect_leaf_uris(rows: list[dict]) -> dict[str, str]:
    """Return a mapping of URI -> short title derived from data rows."""
    uris = {}
    for r in rows:
        uris[r["tgt_uri"]] = "CE.xpt"
        src_uri = r["src_uri"]
        col = r["src_col"].upper()
        if col in ("ICD10", "TERM", "DATE") and "PT_DX" not in uris.get(src_uri, ""):
            # Heuristic: PT_DX if column names match diagnoses table
            uris.setdefault(src_uri, "PT_DX.csv")
        elif col in ("VITAL", "VALUE") or col == "DATE":
            uris.setdefault(src_uri, src_uri.split("/")[-1] + ".csv")
        else:
            uris.setdefault(src_uri, "source.csv")
    return uris


# ---------------------------------------------------------------------------
# RWD-Lineage XML
# ---------------------------------------------------------------------------

def build_rwd_lineage(rows: list[dict]) -> ET.Element:
    ET.register_namespace("", RWDL_NS_URI)
    ET.register_namespace("xsi", XSI_NS)

    root = ET.Element(
        "RWDLineage",
        attrib={
            "xmlns": RWDL_NS_URI,
            "xmlns:xsi": XSI_NS,
            "xsi:schemaLocation": f"{RWDL_NS_URI} rwd-lineage.xsd",
            "FileType": "Snapshot",
            "FileOID": "RWDLineage.Example1",
            "CreationDateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "AsOfDateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "Originator": "CDISC RWD Lineage Project",
            "SourceSystem": "Example",
            "SourceSystemVersion": "1.0",
        },
    )

    for r in rows:
        has_span = r["span_begin"] is not None
        transform_type = "NLPExtraction" if has_span else "DirectMap"
        transform_desc = "NLP Extraction" if has_span else "Direct Map"

        src_coords = f"{r['src_uri']}|{r['src_row']}|{r['src_col']}"
        tgt_coords = f"{r['tgt_uri']}|{r['tgt_row']}|{r['tgt_col']}"
        uid = make_uuid(src_coords, tgt_coords)

        m = ET.SubElement(root, "MapID", attrib={"uuid": uid})
        t = ET.SubElement(m, "Transformation", attrib={"type": transform_type})
        t.text = transform_desc

        src_el = ET.SubElement(m, "Source")
        src_coord = ET.SubElement(
            src_el, "Coordinate",
            attrib={"storage": r["src_storage"], "structure": r["src_structure"]},
        )
        ET.SubElement(src_coord, "URI").text = r["src_uri"]
        ET.SubElement(src_coord, "RowIndex").text = str(r["src_row"])
        ET.SubElement(src_coord, "ColumnName").text = r["src_col"]
        if has_span:
            ET.SubElement(
                src_coord, "TextSpan",
                attrib={"begin": str(r["span_begin"]), "end": str(r["span_end"])},
            )

        tgt_el = ET.SubElement(m, "Target")
        tgt_coord = ET.SubElement(
            tgt_el, "Coordinate",
            attrib={"storage": r["tgt_storage"], "structure": r["tgt_structure"]},
        )
        ET.SubElement(tgt_coord, "URI").text = r["tgt_uri"]
        ET.SubElement(tgt_coord, "RowIndex").text = str(r["tgt_row"])
        ET.SubElement(tgt_coord, "ColumnName").text = r["tgt_col"]

    print(f"Generated {len(rows)} MapID elements")
    return root


# ---------------------------------------------------------------------------
# Define-XML
# ---------------------------------------------------------------------------

def build_define_xml(rows: list[dict]) -> ET.Element:
    ET.register_namespace("",      DEFINE_NS)
    ET.register_namespace("def",   DEFINE_NS2)
    ET.register_namespace("rwdl",  RWDL_NS_URI)
    ET.register_namespace("xsi",   XSI_NS)
    ET.register_namespace("xlink", XL_NS)

    D = DEFINE_NS2

    root = ET.Element(
        f"{{{DEFINE_NS}}}ODM",
        attrib={
            f"{{{XSI_NS}}}schemaLocation": (
                f"{DEFINE_NS} "
                "https://www.cdisc.org/standards/odm/v1.3.2/ODM1-3-2.xsd"
            ),
            "FileType": "Snapshot",
            "FileOID": "Define.Example1",
            "CreationDateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "AsOfDateTime": datetime.now().strftime("%Y-%m-%dT%H:%M:%S"),
            "Originator": "CDISC RWD Lineage Project",
            "SourceSystem": "Example",
            "SourceSystemVersion": "1.0",
            f"{{{D}}}Context": "Submission",
        },
    )

    study = ET.SubElement(root, f"{{{DEFINE_NS}}}Study", attrib={"OID": "Study.Example1"})
    gv = ET.SubElement(study, f"{{{DEFINE_NS}}}GlobalVariables")
    ET.SubElement(gv, f"{{{DEFINE_NS}}}StudyName").text = "RWD Lineage Example 1"
    ET.SubElement(gv, f"{{{DEFINE_NS}}}StudyDescription").text = (
        "Example 1: CE domain with EHR source lineage"
    )
    ET.SubElement(gv, f"{{{DEFINE_NS}}}ProtocolName").text = "RWDL-EX1"

    mdv = ET.SubElement(
        study,
        f"{{{DEFINE_NS}}}MetaDataVersion",
        attrib={
            "OID": "MDV.Example1",
            "Name": "Example 1 Define-XML",
            f"{{{D}}}DefineVersion": "2.1.0",
            f"{{{D}}}StandardName": "SDTM",
            f"{{{D}}}StandardVersion": "1.9",
        },
    )

    # -- Annotations / lineage reference --
    ann = ET.SubElement(mdv, f"{{{D}}}Annotations")
    lineage_el = ET.SubElement(ann, f"{{{RWDL_NS_URI}}}lineage")
    ref_el = ET.SubElement(
        lineage_el, f"{{{RWDL_NS_URI}}}ref", attrib={"leafID": "LF.RWDLINEAGE"}
    )
    ref_el.text = "rwd-lineage.xml"

    # -- ItemGroupDef: CE --
    igdef = ET.SubElement(
        mdv,
        f"{{{DEFINE_NS}}}ItemGroupDef",
        attrib={
            "OID": "IG.CE",
            "Name": "CE",
            "Repeating": "Yes",
            "IsReferenceData": "No",
            "SASDatasetName": "CE",
            f"{{{D}}}Structure": "One record per subject per clinical event",
            f"{{{D}}}Purpose": "Tabulation",
            f"{{{D}}}StandardOID": "STD.1",
            f"{{{D}}}ArchiveLocationID": "LF.CE",
        },
    )
    desc = ET.SubElement(igdef, f"{{{DEFINE_NS}}}Description")
    ET.SubElement(desc, f"{{{DEFINE_NS}}}TranslatedText").text = "Clinical Events"

    for item_oid, mandatory in [
        ("IT.CE.STUDYID", "Yes"),
        ("IT.CE.DOMAIN",  "Yes"),
        ("IT.CE.USUBJID", "Yes"),
        ("IT.CE.CESEQ",   "Yes"),
        ("IT.CE.CETERM",  "Yes"),
        ("IT.CE.CEPRESP", "No"),
        ("IT.CE.CEOCCUR", "No"),
    ]:
        ET.SubElement(
            igdef,
            f"{{{DEFINE_NS}}}ItemRef",
            attrib={"ItemOID": item_oid, "Mandatory": mandatory},
        )

    # -- ItemDefs --
    items = [
        ("IT.CE.STUDYID", "STUDYID", "Study Identifier",         "text",    8),
        ("IT.CE.DOMAIN",  "DOMAIN",  "Domain Abbreviation",      "text",    2),
        ("IT.CE.USUBJID", "USUBJID", "Unique Subject Identifier", "text",   50),
        ("IT.CE.CESEQ",   "CESEQ",   "Sequence Number",           "integer", 8),
        ("IT.CE.CETERM",  "CETERM",  "Reported Term for the CE",  "text",  200),
        ("IT.CE.CEPRESP", "CEPRESP", "CE Pre-Specified",          "text",    1),
        ("IT.CE.CEOCCUR", "CEOCCUR", "CE Occurrence",             "text",    1),
    ]
    for oid, name, label, dtype, length in items:
        idef = ET.SubElement(
            mdv,
            f"{{{DEFINE_NS}}}ItemDef",
            attrib={
                "OID": oid,
                "Name": name,
                "DataType": dtype,
                "Length": str(length),
                f"{{{D}}}Label": label,
            },
        )
        idesc = ET.SubElement(idef, f"{{{DEFINE_NS}}}Description")
        ET.SubElement(idesc, f"{{{DEFINE_NS}}}TranslatedText").text = label

    # -- Leaf elements: derive unique source URIs from lineage rows --
    tgt_uri = rows[0]["tgt_uri"]
    src_uris = list(dict.fromkeys(r["src_uri"] for r in rows))  # ordered, unique

    leaf_titles = {
        tgt_uri: ("LF.CE", "CE.xpt"),
    }
    for src_uri in src_uris:
        # Determine name from column names used with this URI
        cols = {r["src_col"].upper() for r in rows if r["src_uri"] == src_uri}
        if "ICD10" in cols or "TERM" in cols:
            leaf_titles[src_uri] = (f"LF.{src_uri.split('/')[-2][:6].upper()}_DX", "PT_DX.csv")
        elif "VITAL" in cols or "VALUE" in cols:
            leaf_titles[src_uri] = ("LF.VITALS", "VITALS.csv")
        elif "TEXT" in cols:
            leaf_titles[src_uri] = ("LF.NOTES", "NOTES.csv")
        else:
            leaf_titles[src_uri] = ("LF.SOURCE", "source.csv")

    # Fixed leaf IDs based on known URIs for stability
    known_leaf_ids = {
        "1zqMXZviKvS_bcC11IaSsHuZVoOhhD3D7": ("LF.CE",     "CE.xpt"),
        "1SukPxF0W6aPjDMnlluHGxsSiDMq3-J9V": ("LF.PT_DX",  "PT_DX.csv"),
        "1ofgeU1M8rc2bnu0ygR5vwsWE7mRIexFz":  ("LF.VITALS", "VITALS.csv"),
        "1D8zqaMGsmgCWRX8lX4WnIct4Mr2qxVYg":  ("LF.NOTES",  "NOTES.csv"),
    }

    all_uris = [tgt_uri] + src_uris
    seen = set()
    for uri in all_uris:
        if uri in seen:
            continue
        seen.add(uri)
        file_id = uri.split("/d/")[-1].split("/")[0]
        leaf_id, title = known_leaf_ids.get(file_id, ("LF.UNKNOWN", "unknown.csv"))
        leaf = ET.SubElement(
            mdv,
            f"{{{D}}}leaf",
            attrib={"ID": leaf_id, f"{{{XL_NS}}}href": uri},
        )
        ET.SubElement(leaf, f"{{{D}}}title").text = title

    # RWD-Lineage file leaf
    leaf = ET.SubElement(
        mdv,
        f"{{{D}}}leaf",
        attrib={"ID": "LF.RWDLINEAGE", f"{{{XL_NS}}}href": "rwd-lineage.xml"},
    )
    ET.SubElement(leaf, f"{{{D}}}title").text = "rwd-lineage.xml"

    return root


# ---------------------------------------------------------------------------
# CSV exports
# ---------------------------------------------------------------------------

CSV_EXPORTS = {
    "SDTM CE":       Path(__file__).parent / "data" / "csv" / "sdtm"   / "ce.csv",
    "Source PT_DX":  Path(__file__).parent / "data" / "csv" / "source" / "pt_dx.csv",
    "Source VITALS": Path(__file__).parent / "data" / "csv" / "source" / "vitals.csv",
    "Source NOTES":  Path(__file__).parent / "data" / "csv" / "source" / "notes.csv",
}


def export_csvs():
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    for sheet_name, out_path in CSV_EXPORTS.items():
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # Strip trailing all-None rows
        while rows and all(v is None for v in rows[-1]):
            rows.pop()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        with open(out_path, "w", newline="") as f:
            writer = csv.writer(f)
            for row in rows:
                writer.writerow(["" if v is None else v for v in row])
        print(f"Wrote {out_path}  ({len(rows) - 1} data rows)")


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    rows = load_lineage_rows()

    # CSV exports
    export_csvs()

    # RWD-Lineage XML
    lineage_root = build_rwd_lineage(rows)
    lineage_path = OUTPUT_DIR / "rwd-lineage.xml"
    lineage_path.write_text(pretty_xml(lineage_root), encoding="utf-8")
    print(f"Wrote {lineage_path}")

    # Define-XML
    define_root = build_define_xml(rows)
    define_path = OUTPUT_DIR / "define.xml"
    define_path.write_text(pretty_xml(define_root), encoding="utf-8")
    print(f"Wrote {define_path}")


if __name__ == "__main__":
    main()
